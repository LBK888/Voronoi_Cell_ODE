import numpy as np
from scipy.spatial import Voronoi

class VoronoiGrid:
    def __init__(self, grid_shape=(2,1), cell_dist=1.0, pos_rand=0.0, mode='honeycomb', custom_cells=None, import_path=None):
        """
        grid_shape: (x, y) 格狀排列
        cell_dist: 細胞間距
        pos_rand: 位置隨機擾動幅度
        mode: 'honeycomb', 'random', 'regular', 'custom', 'import'
        custom_cells: np.ndarray, 若 mode='custom' 則用此
        import_path: str, 若 mode='import' 則從檔案讀取
        """
        self.grid_shape = grid_shape
        self.cell_dist = cell_dist
        self.pos_rand = pos_rand
        self.mode = mode
        self.custom_cells = custom_cells
        self.import_path = import_path
        self.cells = self._init_cells()
        self.vor = Voronoi(self.cells)

    def _init_cells(self):
        import os
        import numpy as np
        if self.mode == 'honeycomb':
            cells = []
            for x in range(self.grid_shape[0]):
                for y in range(self.grid_shape[1]):
                    if y % 2 == 0:
                        px = x + (1 + (np.random.random_sample() - 0.5) * self.pos_rand) * self.cell_dist
                        py = y + (1 + (np.random.random_sample() - 0.5) * self.pos_rand) * self.cell_dist
                    else:
                        px = x + (1 + (np.random.random_sample() - 0.5) * self.pos_rand) * self.cell_dist + 0.5 * self.cell_dist
                        py = y + (1 + (np.random.random_sample() - 0.5) * self.pos_rand) * self.cell_dist
                    cells.append([px, py])
            return np.array(cells)
        elif self.mode == 'random':
            n = self.grid_shape[0] * self.grid_shape[1]
            cells = np.random.rand(n, 2) * np.array([self.grid_shape[0], self.grid_shape[1]]) * self.cell_dist
            return cells
        elif self.mode == 'regular':
            xs = np.arange(self.grid_shape[0]) * self.cell_dist
            ys = np.arange(self.grid_shape[1]) * self.cell_dist
            grid = np.array(np.meshgrid(xs, ys)).T.reshape(-1, 2)
            return grid
        elif self.mode == 'custom':
            if self.custom_cells is not None:
                return np.array(self.custom_cells)
            else:
                raise ValueError('custom_cells must be provided for custom mode')
        elif self.mode == 'import':
            if self.import_path is not None:
                ext = os.path.splitext(self.import_path)[-1].lower()
                if ext == '.xlsx':
                    try:
                        from openpyxl import load_workbook
                    except ImportError:
                        raise ImportError('openpyxl is required to import xlsx files')
                    wb = load_workbook(self.import_path, read_only=True, data_only=True)
                    ws = wb.active
                    headers = [str(cell.value).strip().lower() if cell.value is not None else '' for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                    # 優先找 T, dT, x, y
                    x_idx = None
                    y_idx = None
                    t_idx = None
                    step_idx = None
                    for idx, h in enumerate(headers):
                        if h == 'x':
                            x_idx = idx
                        elif h == 'y':
                            y_idx = idx
                        elif h == 't':
                            t_idx = idx
                        elif h == 'step':
                            step_idx = idx
                    if x_idx is not None and y_idx is not None:
                        # 若有 T/dT 欄，僅取 T=0/dT=0 的資料
                        rows = list(ws.iter_rows(min_row=2, values_only=True))
                        filtered = []
                        if t_idx is not None and step_idx is not None:
                            for row in rows:
                                if row[t_idx] == 0 and row[step_idx] == 0:      #the t=0 cell grid
                                    filtered.append([row[x_idx], row[y_idx]])
                        else:
                            # 沒有 T/dT 欄，全部取 x, y
                            filtered = [[row[x_idx], row[y_idx]] for row in rows if row[x_idx] is not None and row[y_idx] is not None]
                        if not filtered:
                            raise ValueError('No valid cell positions found in xlsx (T=0/step=0 or all x/y)')
                        return np.array(filtered)
                    else:
                        raise ValueError('xlsx 檔案需包含 x, y 欄位 (不分大小寫)')
                else:
                    # 預設 csv/txt
                    try:
                        return np.loadtxt(self.import_path, delimiter=',')
                    except Exception as e:
                        raise ValueError(f'Failed to load import file: {e}')
            else:
                raise ValueError('import_path must be provided for import mode')
        else:
            raise ValueError(f'Unknown mode: {self.mode}')

    def move_cells(self, rule=None, random_strength=0.0):
        """
        rule: function(cells) -> new_cells
        random_strength: 0~1, 隨機移動強度（相對 cell_dist）
        """
        new_cells = self.cells.copy()
        if rule:
            new_cells = rule(new_cells)
        if random_strength > 0:
            new_cells += (np.random.rand(*new_cells.shape)-0.5)*2*self.cell_dist*random_strength
        self.cells = new_cells
        self.vor = Voronoi(self.cells) 

    def get_inner_cell_indices(self):
        """
        回傳內圈細胞的 index。
        1. 先排除 region 中帶有 -1 的細胞（外圈）
        2. 再排除與外圈細胞相鄰的細胞（利用 ridge_points）
        """
        vor = self.vor
        # 1. 找出所有外圈細胞（region 含 -1）
        outer_set = set()
        for i, region_index in enumerate(vor.point_region):
            region = vor.regions[region_index]
            if not region or any(v < 0 for v in region):
                outer_set.add(i)
        # 2. 找出所有與外圈細胞相鄰的細胞
        neighbor_set = set()
        for p1, p2 in vor.ridge_points:
            if p1 in outer_set:
                neighbor_set.add(p2)
            if p2 in outer_set:
                neighbor_set.add(p1)
        # 3. 內圈細胞 = 全部細胞 - 外圈 - 鄰居
        all_indices = set(range(len(vor.points)))
        inner_indices = list(all_indices - outer_set - neighbor_set)
        return inner_indices

    def cell_proliferation(self, n=1, mode='area', concentrations=None):
        from scipy.spatial import ConvexHull
        new_cells = []
        new_conc = []
        for _ in range(n):
            vor = self.vor  # 每次都用最新的
            inner_indices = self.get_inner_cell_indices()
            # 計算面積
            areas = []
            for i in inner_indices:
                region = vor.regions[vor.point_region[i]]
                polygon = [vor.vertices[v] for v in region]
                try:
                    hull = ConvexHull(polygon)
                    areas.append(hull.volume)
                except:
                    areas.append(0)
            if not areas:
                break
            # 選最大面積
            if mode == 'area':
                idx = inner_indices[np.argmax(areas)]
            elif mode == 'random':
                idx = np.random.choice(inner_indices)
            else:
                raise ValueError('mode 必須為 area 或 random')
            # 分裂
            region = vor.regions[vor.point_region[idx]]
            polygon = np.array([vor.vertices[v] for v in region])
            center = self.cells[idx]
            cov = np.cov(polygon - center, rowvar=False)
            eigvals, eigvecs = np.linalg.eigh(cov)
            axis = eigvecs[:, np.argmax(eigvals)]
            offset = axis * self.cell_dist * 0.75    # 0.75 is a magic number
            self.cells[idx] = self.cells[idx] - offset
            new_cell = center + offset
            self.cells = np.vstack([self.cells, new_cell])
            if concentrations is not None:
                new_conc.append(concentrations[idx])
                concentrations = np.vstack([concentrations, concentrations[idx]])
            self.vor = Voronoi(self.cells)
            new_cells.append(len(self.cells)-1)
        if concentrations is not None:
            return new_cells, concentrations
        return new_cells

    def cell_apoptosis(self, n=1, mode='area', concentrations=None):
        """
        細胞死亡：移除面積最小或隨機選的 n 個細胞
        mode: 'area'（預設，最小面積）或 'random'（隨機）
        concentrations: (cell, var) array，若有提供則同步移除濃度
        回傳：被移除的細胞索引list與新濃度陣列（若有）
        """
        from scipy.spatial import ConvexHull
        cells = self.cells
        vor = self.vor
        # 過濾外圈細胞
        inner_indices = self.get_inner_cell_indices()
        # 計算面積
        areas = []
        for i in inner_indices:
            region = vor.regions[vor.point_region[i]]
            polygon = [vor.vertices[v] for v in region]
            try:
                hull = ConvexHull(polygon)
                areas.append(hull.volume)
            except:
                areas.append(1e9)
        # 選擇要移除的細胞
        if mode == 'area':
            chosen = np.argsort(areas)[:n]
            chosen_indices = [inner_indices[i] for i in chosen]
        elif mode == 'random':
            chosen_indices = list(np.random.choice(inner_indices, n, replace=False))
        else:
            raise ValueError('mode 必須為 area 或 random')
        keep_mask = np.ones(len(cells), dtype=bool)
        keep_mask[chosen_indices] = False
        self.cells = self.cells[keep_mask]
        self.vor = Voronoi(self.cells)
        if concentrations is not None:
            new_conc_arr = concentrations[keep_mask]
            return chosen_indices, new_conc_arr
        return chosen_indices
